from openpyxl import load_workbook


def agregar(ruta: int, datos: dict):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel["Datos del crud"]
    Hoja_datos = Hoja_datos["A2":"F" + str(Hoja_datos.max_row + 1)]
    hoja = Archivo_Exccel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInalizada = 6

    for i in Hoja_datos:
        if not (isinstance(i[0].value, int)):
            identificador = i[0].row
            hoja.cell(row=identificador, column=1).value = identificador - 1
            hoja.cell(row=identificador, column=titulo).value = datos["titulo"]
            hoja.cell(row=identificador, column=descripcion).value = datos["descripcion"]
            hoja.cell(row=identificador, column=estado).value = datos["estado"]
            hoja.cell(row=identificador, column=fecha_inicio).value = datos["fecha inicio"]
            hoja.cell(row=identificador, column=fecha_FInalizada).value = datos["fecha finalizacion"]
            break
    Archivo_Exccel.save(ruta)
    return
