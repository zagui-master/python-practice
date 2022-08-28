from openpyxl import load_workbook


def actualizar(ruta: str, identificador: int, datos_actualizados: dict):
    Archivo_exccel = load_workbook(ruta)
    Hoja_datos = Archivo_exccel["Datos del crud"]
    Hoja_datos = Hoja_datos["A2": "F" + str(Hoja_datos.max_row)]
    hoja = Archivo_exccel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInalizado = 6
    encontro = False

    for i in Hoja_datos:
        if i[0].value == identificador:
            fila = i[0].row
            encontro = True
            for d in datos_actualizados:
                if d == "titulo" and not (datos_actualizados[d] == ""):
                    hoja.cell(row=fila, column=titulo).value = datos_actualizados[d]
                elif d == "descripcion" and not (datos_actualizados[d] == ""):
                    hoja.cell(row=fila, column=descripcion).value = datos_actualizados[d]
                elif d == "estado" and not (datos_actualizados[d] == ""):
                    hoja.cell(row=fila, column=estado).value = datos_actualizados[d]
                elif d == "fecha inicio" and not (datos_actualizados[d] == ""):
                    hoja.cell(row=fila, column=fecha_inicio).value = datos_actualizados[d]
                elif d == "fecha finalizacion" and not (datos_actualizados[d] == ""):
                    hoja.cell(row=fila, column=fecha_FInalizado).value = datos_actualizados[d]

    Archivo_exccel.save(ruta)
    if not encontro:
        print("Error : No existe una tarea con ese Id")
        print()
    return
