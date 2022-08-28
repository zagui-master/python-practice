from openpyxl import load_workbook
from filterItem import filtrar


def leer(ruta: str, extraer: str):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel["Datos del crud"]
    Hoja_datos = Hoja_datos["A2":"F" + str(Hoja_datos.max_row)]

    info = {}

    for i in Hoja_datos:
        if isinstance(i[0].value, int):
            info.setdefault(i[0].value, {"tarea": i[1].value,
                                         "descripcion": i[2].value,
                                         "estado": i[3].value,
                                         "fecha de inicio": i[4].value,
                                         "fecha de finalizacion": i[5].value})

    if not (extraer == "todo"):
        info = filtrar(info, extraer)

    for i in info:
        print('******** Tarea ********')
        print('Id:' + str(i) + "\n" +
              'Titulo :' + str(info[i]["tarea"]) + "\n" +
              'Descripcion :' + str(info[i]["descripcion"]) + "\n" +
              'Estado :' + str(info[i]["estado"]) + "\n" +
              'Fecha creacion :' + str(info[i]["fecha de inicio"]) + "\n" +
              'Fecha de finalizacion :' + str(info[i]["fecha de finalizacion"]) + "\n")
        print()
    return
