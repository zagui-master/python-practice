from openpyxl import load_workbook


def borrar(ruta, identificador):
    Archivo_Exccel = load_workbook(ruta)
    Hoja_datos = Archivo_Exccel["Datos del crud"]
    Hoja_datos = Hoja_datos["A2":"F" + str(Hoja_datos.max_row + 1)]
    hoja = Archivo_Exccel.active

    titulo = 2
    descripcion = 3
    estado = 4
    fecha_inicio = 5
    fecha_FInal = 6
    encontrado = False

    for i in Hoja_datos:
        if i[0].value == identificador:
            fila = i[0].row
            encontrado = True
            hoja.cell(row=fila, column=1).value = ""
            hoja.cell(row=fila, column=titulo).value = ""
            hoja.cell(row=fila, column=descripcion).value = ""
            hoja.cell(row=fila, column=estado).value = ""
            hoja.cell(row=fila, column=fecha_inicio).value = ""
            hoja.cell(row=fila, column=fecha_FInal).value = ""

    Archivo_Exccel.save(ruta)
    if not encontrado:
        print("Error: No existe una tarea con es Id")
        print()
    return